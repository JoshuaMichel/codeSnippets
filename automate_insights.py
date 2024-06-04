# Given a path to an excel document of data, performs a series of transformations 
# specifically in excel and saves the file in the specified output
# Used Tkinter library to add a GUI user interface for user to 
# change the default input and output paths

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
import numpy as np
from datetime import datetime, timedelta

# Constants for Style Attributes
EVEN_ROW_FILL_COLOR = 'FFD9EAD3'
HEADER_FILL_COLOR = 'FF00B050'
HEADER_TEXT_COLOR = 'FFFFFF'
HEADER_BOLD = True
OUTPUT_SHEET_NAME = "insights"

def preprocess_dataframe(df):
    # Perform preprocessing on the DataFrame
    df = df.replace({pd.NaT: '', np.nan: ''})

    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col], format='%b %d %Y %I:%M:%S %p')
                df[col] = df[col].dt.strftime('%m/%d/%Y %H:%M:%S')
            except ValueError:
                pass
        elif pd.api.types.is_object_dtype(df[col]):
            df[col] = df[col].str.upper()

    df['Email'] = df['Email'].replace({pd.NaT: '', np.nan: ''})
    return df

def set_excel_data_types(df):
    # Specify data types for Excel columns
    excel_dtypes = {
        'Agent Name': 'str',
        'User Query': 'str',
        'Language': 'str',
        'User': 'str',
        'User Id': 'int',
        'Email': 'str',
        'Node': 'str',
        'Skill Key': 'str',
        'Intent Key': 'str',
        'Intent': 'str',
        'Intent Type': 'str',
        'Goal': 'str',
        'Channel': 'str',
        'Featured Tokens': 'str',
        'Entities': 'str',
        'Tags': 'str',
        'Tone': 'str',
        'Sentiment': 'str',
        'Analyzed Document': 'str',
        'Negation': 'bool',
        'Created At': 'datetime64[ns]',
        'Created At (UTC)': 'datetime64[ns]'
    }
    return df.astype(excel_dtypes, errors='ignore')

def add_filters(sheet, df):
    # Add filters to the Excel file
    for idx, column in enumerate(df.columns, start=1):
        col_letter = chr(ord('A') + idx - 1)
        sheet.auto_filter.ref = f'A1:{col_letter}{len(df) + 1}'

def adjust_column_width(sheet, df):
    # Adjust column width for all columns excluding specified columns
    exclude_columns = ['A', 'E', 'F', 'G', 'I', 'K', 'M', 'Q', 'R', 'T']
    for col_num, column in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_num)
        if col_letter not in exclude_columns:
            sheet.column_dimensions[col_letter].width = 20

def set_alignment_and_colors(sheet, df):
    # Set alignment, background colors, and header style
    for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_col=len(df.columns), max_row=len(df) + 1), start=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_col=len(df.columns), max_row=len(df) + 1), start=2):
        if row_num % 2 == 0:
            for cell in row:
                cell.fill = PatternFill(start_color=EVEN_ROW_FILL_COLOR, end_color=EVEN_ROW_FILL_COLOR, fill_type='solid')

    for col_num, cell in enumerate(sheet[1], start=1):
        col_letter = get_column_letter(col_num)
        cell.fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type='solid')
        cell.font = Font(color=HEADER_TEXT_COLOR, bold=HEADER_BOLD)

    sheet[1][0].value = "Bot Name"  # Change the value of the first cell in the header row

def perform_csv_transformation(input_path, output_path):
    # Read CSV and perform initial preprocessing
    df = pd.read_csv(input_path)
    df = preprocess_dataframe(df)
    df = set_excel_data_types(df)

    # Write to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name=OUTPUT_SHEET_NAME, header=True)
        sheet = writer.sheets[OUTPUT_SHEET_NAME]

        # Add filters and adjust column width
        add_filters(sheet, df)
        adjust_column_width(sheet, df)

        # Set alignment and background colors
        set_alignment_and_colors(sheet, df)

        print(f"Transformation complete. Result saved to: {output_path}")

    
def get_last_sunday_date():
    today = datetime.now()
    days_to_sunday = today.weekday() - today.weekday() % 7
    last_sunday = today - timedelta(days=days_to_sunday)
    return last_sunday.strftime('%m%d%y')

if __name__ == "__main__":
    input_file_path = "DEFAULT INPUT PATH HERE"
    output_folder = r"DEFAULT OUTPUT PATH HERE"
    output_file_name = f"insights{get_last_sunday_date()}.xlsx"
    output_file_path = os.path.join(output_folder, output_file_name)

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    perform_csv_transformation(input_file_path, output_file_path)